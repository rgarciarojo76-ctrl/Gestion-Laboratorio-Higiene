import os
import re
import openpyxl
from docx import Document

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
# Note: Ensure the exact path and extension
DOCX_PATH = os.path.join(BASE_DIR, "MTMHI2025v01.docx")
EXCEL_PATH = os.path.join(BASE_DIR, "data", "Base de conocimiento completa 2026.xlsx")

def clean_volume_string(text):
    """
    Applies the Master Prompt rules:
    - DISCARD lines that contain "L/min".
    - CAPTURE lines that contain "L".
    - CRITICAL RULE: If digits immediately follow "L" without space (superscripts), remove them.
      (e.g., '1-15 L37' -> '1-15 L')
    - If space or parenthesis after "L", keep them. (e.g., '8-400 L (120 L)' -> kept)
    """
    # Split the text by lines in the cell
    lines = text.split('\n')
    valid_lines = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Rule 1: Discard if contains 'L/min'
        if 'L/min' in line:
            continue
            
        # Rule 2: Capture if contains 'L'
        if 'L' in line:
            # Rule 3: Strip digits immediately following 'L'
            # regex: look for 'L' followed by 1 or more digits, followed by end of string or whitespace or parenthesis
            # We replace 'L\d+' with 'L'
            cleaned_line = re.sub(r'L(\d+)', r'L', line)
            valid_lines.append(cleaned_line)
            
    if valid_lines:
        return "\n".join(valid_lines)
    return None

def extract_volumes_from_docx(docx_path):
    """
    Reads the MTMHI DOCX tables. Returns a dictionary mapping CAS number -> Cleaned Volume string.
    """
    if not os.path.exists(docx_path):
        print(f"Error: {docx_path} not found.")
        return {}

    doc = Document(docx_path)
    extracted_data = {}
    
    for table in doc.tables:
        if not table.rows:
            continue
            
        # Find headers
        header_row = table.rows[0]
        cas_col_idx = -1
        caudal_v_col_idx = -1
        
        for idx, cell in enumerate(header_row.cells):
            text = cell.text.lower().strip()
            if "cas" in text:
                cas_col_idx = idx
            elif "caudal y v" in text or "caudal y v." in text or "volumen" in text:
                caudal_v_col_idx = idx
                
        if cas_col_idx == -1 or caudal_v_col_idx == -1:
            # Maybe the header is in the second row (e.g., grouped headers)
            if len(table.rows) > 1:
               header_row_2 = table.rows[1]
               for idx, cell in enumerate(header_row_2.cells):
                   text = cell.text.lower().strip()
                   if "cas" in text and cas_col_idx == -1:
                       cas_col_idx = idx
                   elif ("caudal y v" in text or "caudal y v." in text or "volumen" in text) and caudal_v_col_idx == -1:
                       caudal_v_col_idx = idx

        if cas_col_idx == -1 or caudal_v_col_idx == -1:
            continue # Skip tables without these columns
            
        # Parse data rows
        for row in table.rows[1:]:
            try:
                cas_text = row.cells[cas_col_idx].text.strip()
                vol_text = row.cells[caudal_v_col_idx].text.strip()
                
                # Verify it's a valid CAS format (digits-digits-digits)
                if not re.search(r'\d+-\d+-\d+', cas_text):
                    continue
                    
                # Extract the CAS (sometimes there's more text in the cell, we just take the first CAS format found)
                cas_match = re.search(r'\d+-\d+-\d+', cas_text)
                cas_num = cas_match.group(0) if cas_match else None
                
                if cas_num and vol_text:
                    cleaned_vol = clean_volume_string(vol_text)
                    if cleaned_vol:
                        extracted_data[cas_num] = cleaned_vol
            except IndexError:
                continue

    return extracted_data

def update_excel_database(volumes_data):
    """
    Overwrites the 'Volumen Método' column in the Master Knowledgebase Excel.
    """
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Database {EXCEL_PATH} not found.")
        return

    print("Loading Excel Database...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet_name = "VLA-ED (y VLA-EC) pruebas ind."
    if sheet_name not in wb.sheetnames:
        print(f"Error: {sheet_name} not found in Excel.")
        return
        
    ws = wb[sheet_name]
    
    header_row = ws[1]
    cas_idx = -1
    vol_idx = -1
    
    for cell in header_row:
        text = str(cell.value).strip().lower() if cell.value else ""
        if text == "cas":
            cas_idx = cell.col_idx
        elif text.startswith("v mínimo muestreo"):
            # We will use the first 'V mínimo muestreo' column we find
            if vol_idx == -1: 
               vol_idx = cell.col_idx
            
    if cas_idx == -1 or vol_idx == -1:
        print("Error: Could not determine CAS or Volumen columns in Excel headers.")
        # Fallback to known index if search fails, though searching is safer
        cas_idx = 1 # Column A
        vol_idx = 15 # Column O 'Volumen Mínimo (L)' - assuming 1-based index
        print(f"Fallback to indices -> CAS: {cas_idx}, Vol: {vol_idx}")
        
    print(f"Found CAS column index: {cas_idx}, Volume column index: {vol_idx}")

    updated_count = 0
    not_found_in_docx = 0

    for row in ws.iter_rows(min_row=2):
        cas_cell = row[cas_idx - 1]
        vol_cell = row[vol_idx - 1]
        
        cas = str(cas_cell.value).strip() if cas_cell.value else ""
        
        if cas in volumes_data:
            # Overwrite!
            old_val = str(vol_cell.value) if vol_cell.value else ""
            new_val = volumes_data[cas]
            
            if old_val != new_val:
                vol_cell.value = new_val
                updated_count += 1
        elif cas and cas != "None":
            not_found_in_docx += 1

    print(f"Saving Excel file... ({updated_count} records updated)")
    wb.save(EXCEL_PATH)
    print(f"Done ✅. \nStats: \n  Updated: {updated_count}\n  Untouched/Not in docx: {not_found_in_docx}")

if __name__ == "__main__":
    print(f"Starting MTMHI Volume Extraction...")
    data = extract_volumes_from_docx(DOCX_PATH)
    print(f"Extracted {len(data)} unique valid CAS volumes from the DOCX.")
    if data:
        update_excel_database(data)
    else:
        print("No data extracted to update.")
