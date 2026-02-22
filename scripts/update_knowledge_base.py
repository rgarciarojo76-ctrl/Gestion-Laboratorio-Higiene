#!/usr/bin/env python3
"""
ETL Pipeline: Base de conocimiento completa 2026 → contaminantes.json

PRIMARY SOURCE: 'VLA-ED (y VLA-EC) pruebas ind.' (44 cols)
AUXILIARY SOURCES (cross-correlated by CÓDIGO prueba / CAS):
  - Notas LEP 2025 (Frases H, Notas, RD 665, Gestis TWA/STEL/País)
  - Intranet V0 DTI (Contenedor, Transporte, Plazos, Observaciones)
  - Intranet Cargar PRUEBAS (Código soporte, Soporte alternativo, Desc técnica, Rango)
  - Consultas, respuesta (Precios, Código soporte)
  - CMR, Apéndice 1 (Familia, Compatibilidades, Evaluación Apéndice 1)
  - Screenings, resumen (Screening descripción, Condiciones ED/EC)
  - Buscar alternativas (Flag pruebas sin método)
"""
import json
import os
import openpyxl
import re

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BASE_DIR)
EXCEL_PATH = os.path.join(PROJECT_DIR, "data", "Base de conocimiento completa 2026.xlsx")
OUTPUT_PATH = os.path.join(PROJECT_DIR, "public", "contaminantes.json")


def get_col_index(headers, possible_names):
    """Find column index by matching header name (case-insensitive, newline-agnostic)."""
    if headers is None:
        return None
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_clean = re.sub(r'\s+', ' ', str(h).strip().lower())
        for name in possible_names:
            if name.lower() in h_clean:
                return i
    return None


def safe_str(row, idx):
    """Safely extract a string value from a row at a given index."""
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    val = str(row[idx]).strip()
    if val in ("#DIV/0!", "#REF!", "#N/A", "#VALUE!", "None"):
        return ""
    return val


def safe_num(row, idx):
    """Safely extract a numeric value from a row, returning string or empty."""
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    val = row[idx]
    if isinstance(val, (int, float)):
        # Round to reasonable precision
        if val == int(val):
            return str(int(val))
        return str(round(val, 4))
    s = str(val).strip()
    if s in ("#DIV/0!", "#REF!", "#N/A", "#VALUE!", "None"):
        return ""
    return s


def clean_string(s):
    """Removes inherited Excel artifacts like '(CAS: 123-45-6)' from strings."""
    if not s:
        return ""
    s = re.sub(r'\(CAS:\s*[\d\-]+\)', '', s)
    s = re.sub(r'\s*;\s*$', '', s)
    return s.strip()


def build_lookup(rows, headers, key_names, extract_fields):
    """Build a dict lookup from a sheet, keyed by the first matching key column."""
    key_idx = get_col_index(headers, key_names)
    if key_idx is None:
        return {}

    result = {}
    for row in rows:
        key_val = safe_str(row, key_idx)
        if not key_val:
            continue
        data = {}
        for json_key, col_names in extract_fields.items():
            idx = get_col_index(headers, col_names)
            data[json_key] = safe_str(row, idx)
        result[key_val] = data
    return result


def main():
    print(f"Reading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # =========================================================================
    # 1. PRIMARY SOURCE: 'VLA-ED (y VLA-EC) pruebas ind.' (44 cols)
    # =========================================================================
    ws_master = wb['VLA-ED (y VLA-EC) pruebas ind.']
    m_rows = list(ws_master.iter_rows(values_only=True))
    m_headers = m_rows[0]
    m_data = m_rows[1:]
    print(f"  Master sheet: {len(m_data)} rows x {len(m_headers)} cols")

    # Map all master columns
    M = {
        "contaminante":       ["descr conc"],
        "metodo_analisis":    ["método"],
        "sinonimo":           ["sinónimos"],
        "compuestos":         ["compuestos"],
        "cas":                ["cas"],
        "contaminante_display": ["consulta pruebas activas"],
        "codigo_prueba":      ["código prueba", "código\nprueba"],
        "conc_fact_prueba":   ["conc fact prueba", "conc fact\nprueba"],
        "laboratorio":        ["laboratorio"],
        "tabla":              ["tabla"],
        "codigo_perfil":      ["código perfil", "código\nperfil"],
        "conc_fact_perfil":   ["conc fact perfil", "conc fact\nperfil"],
        "caudal_metodo":      ["caudal método"],
        "caudal_muestreador": ["caudal muestreador"],
        "metodo_interno_basado_en": ["método interno basado en"],
        "caudal_preferente":  ["caudal preferente"],
        "caudal_alternativo": ["caudal alternativo"],
        "lq":                 ["lq"],
        "vla_ed":             ["vla-ed"],
        "vla_ec":             ["vla-ec"],
        "notas_lep":          ["notas en lep"],
        "frases_h":           ["frases h"],
        "rd_665":             ["rd 665"],
        "gestis_twa":         ["gestis twa"],
        "gestis_stel":        ["gestis stel"],
        "gestis_pais":        ["gestis país", "gestis\npaís"],
        "caudal_asignado":    ["caudal asignado"],
        "tiempo_minimo_asignado": ["tiempo mínimo asignado"],
        "tiempo_minimo_muestreo_ed": ["tiempo mínimo muestreo (10%vla-ed)"],
        "tiempo_minimo_muestreo_twa": ["tiempo mínimo muestreo (10%twa)"],
        "v_minimo_muestreo_ed": ["v mínimo muestreo"],  # first match = VLA-ED based
        "loq_concentracion":  ["concentración límite"],
        "v_maximo_muestreo":  ["v máx. muestreo"],
        "ie_minimo_teorico_ed": ["ie mínimo teórico con vla-ed"],
        "ie_minimo_teorico_twa": ["ie mínimo teórico con twa", "ie mínimo teórico\ncon twa"],
        "comentarios_une_689": ["muestreo mínimo de 2h", "muestreo ed, general 2h"],
        "comentarios_generales": ["comentarios"],
    }

    m_idx = {}
    for key, possible in M.items():
        m_idx[key] = get_col_index(m_headers, possible)

    # Special: there are two 'V mínimo muestreo' columns (VLA-ED and TWA)
    # and two IE límite columns. Find them by position.
    # V mínimo muestreo (L) VLA-ED is col 31, TWA is col 32
    # IE límite teórico condiciones VLA-ED and TWA
    for i, h in enumerate(m_headers):
        if h is None:
            continue
        h_clean = str(h).strip().lower().replace("\n", " ")
        if "v mínimo muestreo" in h_clean and "twa" in h_clean:
            m_idx["v_minimo_muestreo_twa"] = i
        if "ie límite teórico" in h_clean and "vlaed" in h_clean.replace("-", "").replace(" ", ""):
            m_idx["ie_limite_condiciones_ed"] = i
        if "ie límite teórico" in h_clean and "twa" in h_clean:
            m_idx["ie_limite_condiciones_twa"] = i

    # =========================================================================
    # 2. AUXILIARY: Notas LEP 2025 (indexed by CAS)
    # =========================================================================
    notas_by_cas = {}
    try:
        ws_notas = wb['Notas LEP 2025']
        n_rows = list(ws_notas.iter_rows(values_only=True))
        n_headers = n_rows[0]
        notas_by_cas = build_lookup(n_rows[1:], n_headers,
            ["cas"],
            {
                "frases_h": ["frases h"],
                "notas_lep": ["notas en lep"],
                "rd_665": ["rd 665"],
                "gestis_twa": ["gestis twa"],
                "gestis_stel": ["gestis stel"],
                "gestis_pais": ["gestis país", "gestis\npaís"],
                "vla_ed_2025": ["vla-ed", "vla-ed \n(mg/m3)", "vla-ed\n(mg/m3)"],
                "vla_ec_2025": ["vla-ec", "vla-ec \n(mg/m3)", "vla-ec\n(mg/m3)"],
            }
        )
        print(f"  Notas LEP: {len(notas_by_cas)} entries by CAS")
    except Exception as e:
        print(f"  Warning: Notas LEP: {e}")

    # =========================================================================
    # 3. AUXILIARY: Intranet V0 DTI (indexed by CÓDIGO)
    # =========================================================================
    intra_by_codigo = {}
    try:
        ws_intra = wb['Intranet V0 DTI']
        i_rows = list(ws_intra.iter_rows(values_only=True))
        i_headers = i_rows[0]
        intra_by_codigo = build_lookup(i_rows[1:], i_headers,
            ["código"],
            {
                "contenedor": ["contenedor"],
                "transporte": ["transporte"],
                "plazo_entrega": ["plazo entrega"],
                "observaciones_concepto": ["observaciones concepto"],
                "comentarios_prueba": ["comentarios prueba"],
            }
        )
        print(f"  Intranet V0: {len(intra_by_codigo)} entries by CÓDIGO")
    except Exception as e:
        print(f"  Warning: Intranet V0: {e}")

    # =========================================================================
    # 4. AUXILIARY: Intranet Cargar PRUEBAS (indexed by CÓDIGO prueba)
    # =========================================================================
    intra_pruebas_by_codigo = {}
    try:
        ws_ip = wb['Intranet Cargar PRUEBAS']
        ip_rows = list(ws_ip.iter_rows(values_only=True))
        ip_headers = ip_rows[0]
        intra_pruebas_by_codigo = build_lookup(ip_rows[1:], ip_headers,
            ["código prueba", "código\nprueba"],
            {
                "codigo_soporte": ["código soporte"],
                "codigo_soporte_alt": ["código soporte alternativo"],
                "descripcion_tecnica": ["descripción técnica"],
                "rango_trabajo": ["rango trabajo"],
            }
        )
        print(f"  Intranet PRUEBAS: {len(intra_pruebas_by_codigo)} entries")
    except Exception as e:
        print(f"  Warning: Intranet PRUEBAS: {e}")

    # =========================================================================
    # 5. AUXILIARY: Consultas, respuesta (indexed by CÓDIGO prueba)
    # =========================================================================
    consultas_by_codigo = {}
    try:
        ws_cons = wb['Consultas, respuesta']
        c_rows = list(ws_cons.iter_rows(values_only=True))
        c_headers = c_rows[0]
        consultas_by_codigo = build_lookup(c_rows[1:], c_headers,
            ["código prueba", "código  prueba"],
            {
                "precio_soporte": ["precio soporte"],
                "precio_analisis": ["precio análisis"],
                "codigo_soporte_consulta": ["código soporte"],
                "caudal_rango_metodo": ["caudal, rango método"],
            }
        )
        print(f"  Consultas: {len(consultas_by_codigo)} entries")
    except Exception as e:
        print(f"  Warning: Consultas: {e}")

    # =========================================================================
    # 6. AUXILIARY: CMR, Apéndice 1 (indexed by CÓDIGO prueba)
    # =========================================================================
    cmr_by_codigo = {}
    try:
        ws_cmr = wb['CMR, Apéndice 1']
        cmr_rows = list(ws_cmr.iter_rows(values_only=True))
        cmr_headers = cmr_rows[0]
        cmr_lookup = build_lookup(cmr_rows[1:], cmr_headers,
            ["código prueba", "código  prueba"],
            {
                "familia_cmr": ["familia"],
                "compatibilidades_cmr": ["compatibilidades"],
                "evaluacion_apendice_1": ["comentarios al muestreo según apéndice 1"],
                "concentracion_limite_cmr": ["concentración límite asociada"],
                "cumple_003_vlaed": ["¿cumple ≤0,03"],
                "cumple_v_max": ["¿cumple volumen máximo"],
                "contribucion_exterior": ["¿procede considerar"],
            }
        )
        # Mark all CMR entries
        for k, v in cmr_lookup.items():
            v["is_cmr"] = True
            cmr_by_codigo[k] = v
        print(f"  CMR: {len(cmr_by_codigo)} entries")
    except Exception as e:
        print(f"  Warning: CMR: {e}")

    # =========================================================================
    # 7. AUXILIARY: Screenings, resumen (indexed by CÓDIGO prueba)
    # =========================================================================
    screenings_by_codigo = {}
    try:
        ws_scr = wb['Screenings, resumen']
        scr_rows = list(ws_scr.iter_rows(values_only=True))
        scr_headers = scr_rows[0]

        scr_codigo_idx = get_col_index(scr_headers, ["código perfil", "código\nperfil"])
        scr_perfil_idx = get_col_index(scr_headers, ["código perfil", "código\nperfil"])
        scr_desc_idx = get_col_index(scr_headers, ["screening", "descripción"])
        scr_cond_ed_idx = get_col_index(scr_headers, ["condiciones de muestreo como screening para vla-ed"])
        scr_cond_ec_idx = get_col_index(scr_headers, ["condiciones de muestreo como screening para vla-ec"])
        scr_comment_idx = get_col_index(scr_headers, ["comentarios"])

        for row in scr_rows[1:]:
            cod = safe_str(row, scr_codigo_idx)
            if not cod:
                continue
            screenings_by_codigo[cod] = {
                "screening_perfil": safe_str(row, scr_perfil_idx),
                "screening_desc": safe_str(row, scr_desc_idx),
                "screening_condiciones_ed": safe_str(row, scr_cond_ed_idx),
                "screening_condiciones_ec": safe_str(row, scr_cond_ec_idx),
                "screening_comentarios": safe_str(row, scr_comment_idx),
            }
        print(f"  Screenings: {len(screenings_by_codigo)} entries")
    except Exception as e:
        print(f"  Warning: Screenings: {e}")

    # =========================================================================
    # 8. AUXILIARY: Buscar alternativas (flag set)
    # =========================================================================
    alternativas_set = set()
    try:
        ws_alt = wb['Buscar alternativas']
        alt_rows = list(ws_alt.iter_rows(values_only=True))
        
        # Find header row containing 'SUSTANCIA'
        alt_header_row = None
        for row in alt_rows[:10]:
            if row and "SUSTANCIA" in [str(x).strip().upper() for x in row if x]:
                alt_header_row = row
                break
                
        if alt_header_row:
            alt_sustancia_idx = get_col_index(alt_header_row, ["sustancia"])
            if alt_sustancia_idx is not None:
                # Add substances, splitting by newlines, slashes, or commas
                for row in alt_rows:
                    cell = safe_str(row, alt_sustancia_idx)
                    if cell and cell.lower() != "sustancia":
                        # Split by common delimiters in that column
                        import re
                        tokens = re.split(r'[\n/,]+', cell)
                        for token in tokens:
                            t_clean = token.strip().lower()
                            if t_clean:
                                alternativas_set.add(t_clean)
        
        print(f"  Buscar alternativas: {len(alternativas_set)} entries")
    except Exception as e:
        print(f"  Warning: Buscar alternativas: {e}")

    # =========================================================================
    # BUILD MAIN DATASET
    # =========================================================================
    contaminants = []
    seen_ids = set()

    for row in m_data:
        if not any(row):
            continue

        c = {}

        # --- Core fields from master sheet ---
        for json_key, idx in m_idx.items():
            c[json_key] = safe_str(row, idx)

        # Clean CAS and sinonimo
        c["sinonimo"] = clean_string(c.get("sinonimo", ""))
        c["compuestos"] = clean_string(c.get("compuestos", ""))

        # Fallback: if sinonimo empty, use compuestos
        if not c.get("sinonimo") and c.get("compuestos"):
            c["sinonimo"] = c["compuestos"]

        # Numeric fields (prefer numeric extraction)
        for num_key in ["lq", "vla_ed", "vla_ec", "caudal_metodo", "caudal_muestreador",
                        "caudal_preferente", "caudal_alternativo", "caudal_asignado",
                        "tiempo_minimo_asignado", "tiempo_minimo_muestreo_ed",
                        "tiempo_minimo_muestreo_twa", "loq_concentracion",
                        "v_maximo_muestreo", "ie_minimo_teorico_ed", "ie_minimo_teorico_twa",
                        "gestis_twa", "gestis_stel"]:
            idx = m_idx.get(num_key)
            if idx is not None:
                c[num_key] = safe_num(row, idx)

        # V mínimo muestreo
        for extra_key in ["v_minimo_muestreo_ed", "v_minimo_muestreo_twa",
                          "ie_limite_condiciones_ed", "ie_limite_condiciones_twa"]:
            idx = m_idx.get(extra_key)
            if idx is not None:
                c[extra_key] = safe_num(row, idx)

        # --- Build unique ID ---
        codigo = c.get("codigo_prueba", "")
        cas = c.get("cas", "")
        conc_fact = c.get("conc_fact_prueba", "")
        c["id"] = f"{cas or 'NOCAS'}-{codigo or 'NOCODE'}-{conc_fact or 'NOFACT'}"

        # Skip exact duplicates
        if c["id"] in seen_ids:
            continue
        seen_ids.add(c["id"])

        # --- Derive display fields ---
        # Use 'metodo_interno_basado_en' as 'tecnica_analitica'
        c["tecnica_analitica"] = c.get("metodo_interno_basado_en", "") or c.get("metodo_analisis", "")
        c["ref_tecnica"] = c.get("conc_fact_prueba", "")

        # Derived caudal for backward compat
        c["caudal"] = c.get("caudal_asignado", "") or c.get("caudal_preferente", "") or c.get("caudal_metodo", "")
        c["volumen_minimo"] = c.get("v_minimo_muestreo_ed", "")

        # --- Cross-correlate Notas LEP by CAS ---
        if cas and cas in notas_by_cas:
            notas = notas_by_cas[cas]
            
            # Authoritative overwrite: always use 2025 VLAs if available
            if "vla_ed_2025" in notas and str(notas["vla_ed_2025"]).strip():
                val = str(notas["vla_ed_2025"]).strip()
                c["vla_ed"] = val if val.lower() not in ("none", "null", "") else ""
            if "vla_ec_2025" in notas and str(notas["vla_ec_2025"]).strip():
                val = str(notas["vla_ec_2025"]).strip()
                c["vla_ec"] = val if val.lower() not in ("none", "null", "") else ""

            # Only overwrite if master is empty
            for field in ["frases_h", "notas_lep", "rd_665", "gestis_twa", "gestis_stel", "gestis_pais"]:
                if not c.get(field) and notas.get(field):
                    c[field] = notas[field]

        # --- Cross-correlate Intranet V0 DTI by CÓDIGO prueba ---
        if codigo and codigo in intra_by_codigo:
            intra = intra_by_codigo[codigo]
            for field in ["contenedor", "transporte", "plazo_entrega",
                          "observaciones_concepto", "comentarios_prueba"]:
                c[field] = intra.get(field, "")
            # Use Contenedor as soporte_captacion_display
            c["soporte_captacion_display"] = intra.get("contenedor", "") or codigo
        else:
            c["contenedor"] = ""
            c["transporte"] = ""
            c["plazo_entrega"] = ""
            c["observaciones_concepto"] = ""
            c["comentarios_prueba"] = ""
            c["soporte_captacion_display"] = codigo

        c["soporte_captacion"] = codigo
        c["ref_soporte"] = codigo

        # --- Cross-correlate Intranet Cargar PRUEBAS ---
        if codigo and codigo in intra_pruebas_by_codigo:
            ip_data = intra_pruebas_by_codigo[codigo]
            c["codigo_soporte"] = ip_data.get("codigo_soporte", "")
            c["codigo_soporte_alt"] = ip_data.get("codigo_soporte_alt", "")
            c["descripcion_tecnica"] = ip_data.get("descripcion_tecnica", "")
            c["rango_trabajo"] = ip_data.get("rango_trabajo", "")
            # Better soporte display
            if ip_data.get("codigo_soporte"):
                c["ref_soporte"] = ip_data["codigo_soporte"]
            if c.get("soporte_captacion_display") == codigo and ip_data.get("descripcion_tecnica"):
                c["soporte_captacion_display"] = ip_data["descripcion_tecnica"]
        else:
            c["codigo_soporte"] = ""
            c["codigo_soporte_alt"] = ""
            c["descripcion_tecnica"] = ""
            c["rango_trabajo"] = ""

        # --- Cross-correlate Consultas, respuesta ---
        if codigo and codigo in consultas_by_codigo:
            cons = consultas_by_codigo[codigo]
            c["precio_soporte"] = cons.get("precio_soporte", "")
            c["precio_analisis"] = cons.get("precio_analisis", "")
            if not c.get("codigo_soporte") and cons.get("codigo_soporte_consulta"):
                c["codigo_soporte"] = cons["codigo_soporte_consulta"]
            if not c.get("caudal") and cons.get("caudal_rango_metodo"):
                c["caudal_rango_metodo"] = cons["caudal_rango_metodo"]
        else:
            c["precio_soporte"] = ""
            c["precio_analisis"] = ""

        # --- Cross-correlate CMR, Apéndice 1 ---
        if codigo and codigo in cmr_by_codigo:
            cmr = cmr_by_codigo[codigo]
            c["is_cmr"] = True
            c["familia_cmr"] = cmr.get("familia_cmr", "")
            c["compatibilidades_cmr"] = cmr.get("compatibilidades_cmr", "")
            c["evaluacion_apendice_1"] = cmr.get("evaluacion_apendice_1", "")
            c["concentracion_limite_cmr"] = cmr.get("concentracion_limite_cmr", "")
            c["cumple_003_vlaed"] = cmr.get("cumple_003_vlaed", "")
            c["cumple_v_max"] = cmr.get("cumple_v_max", "")
            c["contribucion_exterior"] = cmr.get("contribucion_exterior", "")
        else:
            c["is_cmr"] = False
            c["familia_cmr"] = ""
            c["compatibilidades_cmr"] = ""
            c["evaluacion_apendice_1"] = ""
            c["concentracion_limite_cmr"] = ""
            c["cumple_003_vlaed"] = ""
            c["cumple_v_max"] = ""
            c["contribucion_exterior"] = ""

        # --- Cross-correlate Screenings by Código Perfil analítico ---
        perfil_code = c.get("codigo_perfil", "")
        if perfil_code and perfil_code in screenings_by_codigo:
            scr = screenings_by_codigo[perfil_code]
            c.update(scr)
        else:
            c["screening_perfil"] = ""
            c["screening_desc"] = ""
            c["screening_condiciones_ed"] = ""
            c["screening_condiciones_ec"] = ""
            c["screening_comentarios"] = ""

        # --- Cross-correlate Buscar alternativas ---
        c["sin_metodo_disponible"] = False
        nombre_lower = c.get("contaminante", "").strip().lower()
        display_lower = c.get("contaminante_display", "").strip().lower()
        sinonimo_lower = c.get("sinonimo", "").strip().lower()

        # Build a list of searchable names for this contaminant
        search_names = []
        if nombre_lower: search_names.append(nombre_lower)
        if display_lower: search_names.append(display_lower)
        
        # Add individual synonyms (split by semicolon)
        if sinonimo_lower:
            search_names.extend([s.strip() for s in sinonimo_lower.split(";") if s.strip()])

        for name in search_names:
            if name in alternativas_set:
                c["sin_metodo_disponible"] = True
                break

        contaminants.append(c)

    wb.close()

    # --- Preserve existing visible_en_app values ---
    existing_visibility = {}
    if os.path.exists(OUTPUT_PATH):
        try:
            with open(OUTPUT_PATH, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)
            existing_visibility = {
                c.get('id'): c.get('visible_en_app', True)
                for c in existing_data if c.get('id') is not None
            }
            print(f"  Preserved visibility for {len(existing_visibility)} existing records")
        except Exception as e:
            print(f"  Warning: Could not read existing visibility: {e}")

    for c in contaminants:
        cid = c.get('id')
        if cid is not None and cid in existing_visibility:
            c['visible_en_app'] = existing_visibility[cid]
        else:
            c['visible_en_app'] = True  # Default: visible

    print(f"\nExtracted {len(contaminants)} contaminants. Writing to JSON...")

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(contaminants, f, ensure_ascii=False, indent=2)

    print(f"Successfully updated {OUTPUT_PATH}")

    # Print sample entry for verification
    for c in contaminants:
        if "formaldeh" in c.get("contaminante_display", "").lower():
            print(f"\n=== SAMPLE: {c['contaminante_display']} ===")
            for k, v in sorted(c.items()):
                if v and v not in ("False", ""):
                    print(f"  {k}: {str(v)[:80]}")
            break


if __name__ == "__main__":
    main()
