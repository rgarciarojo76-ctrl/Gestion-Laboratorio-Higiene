import json
import os
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "Base de conocimiento completa 2026.xlsx")
PUBLIC_JSON = os.path.join(BASE_DIR, "public", "contaminantes.json")
DATA_JSON = os.path.join(BASE_DIR, "data", "contaminantes.json")

def sync_volumes():
    print("Loading Excel Database...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["VLA-ED (y VLA-EC) pruebas ind."]
    
    header_row = ws[1]
    cas_idx = -1
    vol_idx = -1
    
    for cell in header_row:
        text = str(cell.value).strip().lower() if cell.value else ""
        if text == "cas":
            cas_idx = cell.col_idx
        elif text.startswith("v mínimo muestreo"):
            if vol_idx == -1: 
               vol_idx = cell.col_idx
               
    if cas_idx == -1 or vol_idx == -1:
        print("Error: Could not determine fields in Excel")
        return

    # Build a lookup dictionary
    vol_data = {}
    for row in ws.iter_rows(min_row=2):
        cas_val = str(row[cas_idx-1].value).strip() if row[cas_idx-1].value else None
        vol_val = str(row[vol_idx-1].value).strip() if row[vol_idx-1].value else None
        if cas_val and cas_val != "None":
            vol_data[cas_val] = vol_val

    # Update both JSONs
    for path in [PUBLIC_JSON, DATA_JSON]:
        if not os.path.exists(path): continue
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            
        updated = 0
        for entry in data:
            cas = entry.get("cas", "")
            if cas in vol_data and vol_data[cas] and vol_data[cas] not in ("None", "#DIV/0!"):
                entry["volumen_minimo"] = vol_data[cas]
                updated += 1
                
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"Updated {updated} records in {path}")

if __name__ == "__main__":
    sync_volumes()
