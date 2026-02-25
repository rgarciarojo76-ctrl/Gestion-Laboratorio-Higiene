import openpyxl
wb = openpyxl.load_workbook("/Users/rubengarciarojo/.gemini/antigravity/workspaces/Gestion-Laboratorio-Higiene/data/Base de conocimiento completa 2026.xlsx", data_only=True)
ws = wb["VLA-ED (y VLA-EC) pruebas ind."]
headers = [str(c.value) for c in ws[1]]
formol = None
for row in ws.iter_rows(min_row=2):
    if str(row[4].value).strip() == "50-00-0":
        formol = [str(c.value) for c in row]
        break
if formol:
    for h, v in zip(headers, formol):
        if "volum" in h.lower() or "v_" in h.lower() or "rango" in h.lower() or "metodo" in h.lower() or "método" in h.lower() or "método" in h.lower() or "15" in v:
            print(f"{h}: {v}")
else:
    print("Formaldehído no encontrado.")
