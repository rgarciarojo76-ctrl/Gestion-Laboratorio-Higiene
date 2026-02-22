import openpyxl

file_path = "data/Base de conocimiento completa 2026.xlsx"
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path, data_only=True)

print("Headers in 'Notas LEP 2025':")
headers = [str(x).strip() for x in next(wb['Notas LEP 2025'].iter_rows(values_only=True)) if x]
print(headers)

print("First few rows:")
for i, row in enumerate(wb['Notas LEP 2025'].iter_rows(values_only=True)):
    if i > 5:
        break
    print([str(x) for x in row[:10]])
